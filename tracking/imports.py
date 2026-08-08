"""Durable, row-isolated tracking spreadsheet import workflow."""

from collections import Counter

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from accounts.authorization import require_permission

from .models import Stage, TrackingImportJob, TrackingImportRow, TrackingEvent
from .services import confirm_stage, get_stage_confirmation_preview


HEADER_ALIASES = {
    "tracking_code": {"tracking_code", "tracking code", "کد رهگیری", "کد_رهگیری"},
    "stage": {"stage", "stage_name", "مرحله", "نام مرحله"},
}


def _normalized_header(value):
    return str(value or "").strip().casefold().replace("-", "_")


def _validation_message(error):
    if isinstance(error, ValidationError):
        return " ".join(error.messages)
    return str(error)


@transaction.atomic
def create_tracking_import_job(*, spreadsheet, actor):
    """Persist an upload, then enqueue it only after its transaction commits."""

    require_permission(
        actor=actor,
        permission="tracking.import_tracking_stage_updates",
        error_message="شما اجازهٔ ورود گروهی مراحل از Excel را ندارید.",
    )

    job = TrackingImportJob(
        upload=spreadsheet,
        original_filename=(spreadsheet.name or "tracking-import.xlsx")[:255],
        requested_by=actor,
        status=TrackingImportJob.Status.QUEUED,
    )
    job.full_clean()
    job.save()

    # Import locally to keep Celery task registration separate from the domain.
    from .tasks import process_tracking_import_job_task

    transaction.on_commit(lambda: process_tracking_import_job_task.delay(job.id))
    return job


def _read_worksheet_rows(job):
    """Read only values from the workbook and return normalized row payloads."""

    try:
        from openpyxl import load_workbook
    except ImportError as error:  # Defensive message for a broken deployment.
        raise RuntimeError("کتابخانهٔ پردازش Excel در سرور نصب نشده است.") from error

    job.upload.open("rb")
    try:
        workbook = load_workbook(job.upload, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValidationError("فایل Excel خالی است.")

        header_positions = {}
        for index, header in enumerate(headers):
            normalized = _normalized_header(header)
            for field_name, aliases in HEADER_ALIASES.items():
                if normalized in aliases and field_name not in header_positions:
                    header_positions[field_name] = index

        missing = [field for field in HEADER_ALIASES if field not in header_positions]
        if missing:
            raise ValidationError(
                "ستون‌های الزامی فایل وجود ندارند: " + ", ".join(missing)
            )

        parsed_rows = []
        for row_number, row in enumerate(rows, start=2):
            tracking_code = str(row[header_positions["tracking_code"]] or "").strip()
            stage_name = str(row[header_positions["stage"]] or "").strip()
            if not tracking_code and not stage_name:
                continue
            parsed_rows.append((row_number, tracking_code, stage_name))
        return parsed_rows
    finally:
        job.upload.close()


def _active_stages_by_name():
    grouped = {}
    for stage in Stage.objects.filter(is_active=True).order_by("order", "pk"):
        grouped.setdefault(stage.name.strip().casefold(), []).append(stage)
    return grouped


def _record_row(*, job, row_number, tracking_code, stage_name, outcome, message=""):
    return TrackingImportRow.objects.create(
        job=job,
        row_number=row_number,
        tracking_code=tracking_code[:40],
        stage_name=stage_name[:120],
        outcome=outcome,
        message=message,
    )


def process_tracking_import_job(*, job_id):
    """Process every spreadsheet row independently through shared tracking rules."""

    with transaction.atomic():
        job = TrackingImportJob.objects.select_for_update().select_related(
            "requested_by"
        ).get(pk=job_id)
        if job.status != TrackingImportJob.Status.QUEUED:
            return {"outcome": "ignored", "job_id": job.id}
        job.status = TrackingImportJob.Status.PROCESSING
        job.started_at = timezone.now()
        job.save(update_fields=["status", "started_at"])

    try:
        # Authorization is re-checked when the Worker actually starts.  A
        # manager may have removed a clearance employee's access while the
        # job was waiting in Redis.
        require_permission(
            actor=job.requested_by,
            permission="tracking.import_tracking_stage_updates",
            error_message="اجازهٔ پردازش این ورود گروهی Excel دیگر معتبر نیست.",
        )
        rows = _read_worksheet_rows(job)
        stages_by_name = _active_stages_by_name()
        duplicate_codes = {
            code.casefold()
            for code, count in Counter(
                code.casefold() for _, code, _ in rows if code
            ).items()
            if count > 1
        }
        success_count = 0
        error_count = 0

        for row_number, tracking_code, stage_name in rows:
            try:
                with transaction.atomic():
                    if not tracking_code or not stage_name:
                        raise ValidationError("کد رهگیری و نام مرحله هر دو الزامی هستند.")
                    if tracking_code.casefold() in duplicate_codes:
                        raise ValidationError("این کد رهگیری بیش از یک‌بار در فایل تکرار شده است.")

                    matching_stages = stages_by_name.get(stage_name.casefold(), [])
                    if len(matching_stages) != 1:
                        raise ValidationError("نام مرحلهٔ فعال معتبر یا یکتا نیست.")
                    requested_stage = matching_stages[0]

                    preview = get_stage_confirmation_preview(
                        tracking_code=tracking_code,
                        staff=job.requested_by,
                    )
                    if preview["stage"].pk != requested_stage.pk:
                        raise ValidationError(
                            f"مرحلهٔ مورد انتظار خودرو «{preview['stage'].name}» است."
                        )

                    confirm_stage(
                        car=preview["car"],
                        stage=requested_stage,
                        staff=job.requested_by,
                        source=TrackingEvent.Source.EXCEL_IMPORT,
                    )
                    _record_row(
                        job=job,
                        row_number=row_number,
                        tracking_code=tracking_code,
                        stage_name=stage_name,
                        outcome=TrackingImportRow.Outcome.SUCCESS,
                        message="ورود خودرو به مرحله با موفقیت ثبت شد.",
                    )
                    success_count += 1
            except Exception as error:
                _record_row(
                    job=job,
                    row_number=row_number,
                    tracking_code=tracking_code,
                    stage_name=stage_name,
                    outcome=TrackingImportRow.Outcome.ERROR,
                    message=_validation_message(error)[:2000],
                )
                error_count += 1

        job.refresh_from_db()
        job.total_rows = len(rows)
        job.success_count = success_count
        job.error_count = error_count
        job.status = (
            TrackingImportJob.Status.COMPLETED
            if not error_count
            else TrackingImportJob.Status.COMPLETED_WITH_ERRORS
        )
        job.finished_at = timezone.now()
        job.failure_reason = ""
        job.save(
            update_fields=[
                "total_rows",
                "success_count",
                "error_count",
                "status",
                "finished_at",
                "failure_reason",
            ]
        )
        return {"outcome": job.status, "job_id": job.id}
    except Exception as error:
        job.refresh_from_db()
        job.status = TrackingImportJob.Status.FAILED
        job.failure_reason = _validation_message(error)[:2000]
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "failure_reason", "finished_at"])
        return {"outcome": "failed", "job_id": job.id}
